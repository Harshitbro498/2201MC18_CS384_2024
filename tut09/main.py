import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Get class and exam dates
def fetch_class_dates():
    class_dates = ["06/08/2024", "13/08/2024", "20/08/2024", "27/08/2024", "03/09/2024", "17/09/2024", "01/10/2024"]
    skipped_dates = ["10/09/2024"]
    exam_dates = ["24/09/2024"]
    return class_dates, skipped_dates, exam_dates


# Load student data from file
def load_students(file_name):
    with open(file_name, 'r') as f:
        student_data = {}
        for line in f.readlines():
            roll_number, name = line.strip().split(maxsplit=1)
            student_data[roll_number] = name
    return student_data

# Parse the attendance CSV
def parse_attendance(csv_file, student_dict):
    attendance_df = pd.read_csv(csv_file)
    
    # Clean and prepare the roll number by extracting the first part and dropping invalid rows
    attendance_df['Roll'] = attendance_df['Roll'].apply(lambda x: x.split()[0] if isinstance(x, str) else None)
    attendance_df = attendance_df.dropna(subset=['Roll'])
    
    # Convert the timestamp to date, assuming day-first format
    attendance_df['Timestamp'] = pd.to_datetime(attendance_df['Timestamp'], dayfirst=True)
    attendance_df['Date'] = attendance_df['Timestamp'].dt.strftime('%d/%m/%Y')
    
    # Create a pivot table to aggregate attendance based on roll and date
    attendance_summary = attendance_df.pivot_table(index='Roll', columns='Date', aggfunc='size', fill_value=0)
    
    # Ensure all students are listed, and append their names
    attendance_summary = attendance_summary.reindex(student_dict.keys(), fill_value=0)
    attendance_summary.index = [f"{roll} {student_dict[roll]}" for roll in attendance_summary.index]
    
    return attendance_summary.copy()

# Create Excel file with attendance information
def create_excel(attendance_data, class_dates):
    # Add columns for the given class dates
    for date in class_dates:
        if date not in attendance_data.columns:
            attendance_data[date] = 0
    
    # Reorder columns according to the class dates
    attendance_data = attendance_data[class_dates].copy()
    
    # Add column for the total number of class days
    attendance_data['Class Days Count'] = len(class_dates)
    
    # Add column for total attendance recorded
    attendance_data['Marked Attendance'] = attendance_data.sum(axis=1)
    
    # Add column for maximum possible attendance
    max_attendance = 2 * len(class_dates)
    attendance_data['Allowed Attendance'] = max_attendance
    
    # Calculate the proxy column (difference between expected and marked attendance)
    attendance_data['Extra Marked'] = abs((2 * len(class_dates)) - attendance_data['Marked Attendance'])
    
    # Save initial output to Excel
    attendance_data.to_excel('output_excel.xlsx', index=True)
    
    # Reload the workbook for further formatting
    workbook = load_workbook('output_excel.xlsx')
    worksheet = workbook.active
    
    # Define color fills for attendance status
    absent_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for absent
    partial_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for partial
    full_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for full attendance
    
    # Apply color coding to attendance values
    for row in worksheet.iter_rows(min_row=2, min_col=2, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            if cell.column <= len(class_dates) + 1:  # Apply to relevant columns only
                if cell.value == 0:
                    cell.fill = absent_fill
                elif cell.value == 1:
                    cell.fill = partial_fill
                elif cell.value == 2:
                    cell.fill = full_fill

    # Save the formatted Excel file
    workbook.save('output_excel.xlsx')

# Main function to execute the process
def execute():
    # Step 1: Load student details from file
    students = load_students('stud_list.txt')
    
    # Step 2: Get relevant class and exam dates
    class_dates, missed_dates, exam_dates = fetch_class_dates()
    
    # Step 3: Parse attendance CSV
    attendance_summary = parse_attendance('input_attendance.csv', students)
    
    # Step 4: Generate Excel report with required formatting
    create_excel(attendance_summary, class_dates)
    print("Excel file with attendance records successfully created: output_excel.xlsx")

if __name__ == '__main__':
    execute()
